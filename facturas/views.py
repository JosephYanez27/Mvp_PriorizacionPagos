"""
views.py — Treasury Prioritization System
All endpoints needed for Screen 1 (Priorización) and Screen 2 (Confirmación/Conciliación).
"""

import json
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction

from .models import Factura, Proveedor


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

TODAY = date(2026, 5, 20)  # Simulated "today" per project requirement


def _serialize_factura(f: Factura) -> dict:
    """Full serialization of a single Factura for the frontend."""
    return {
        "id":                  f.id_factura,
        "folio":               f.folio,
        "proveedor":           f.proveedor.nombre,
        "proveedor_id":        f.proveedor.id_proveedor,
        "tipo_pago":           f.tipo_pago,
        "moneda":              f.moneda,
        "monto_original":      float(f.monto_original),
        "monto_mxn":           float(f.monto_mxn),
        "tipo_cambio":         float(f.tipo_cambio),
        "total":               float(f.total),
        "total_abonado":       float(f.total_abonado),
        "saldo_neto":          float(f.saldo_neto),
        "estado":              f.estado,
        "dias_vencidos":       f.dias_vencidos,
        "scoring":             f.scoring(),
        "estrellas":           float(f.proveedor.prioridad_estrellas),
        "notas_auditoria":     f.notas_auditoria or "",
        "fecha_factura":       f.fecha_factura.isoformat() if f.fecha_factura else None,
        "fecha_registro":      f.fecha_registro.isoformat(),
    }


def _json_error(message: str, status: int = 400) -> JsonResponse:
    return JsonResponse({"ok": False, "error": message}, status=status)


def _json_ok(data=None, **kwargs) -> JsonResponse:
    payload = {"ok": True}
    if data is not None:
        payload["data"] = data
    payload.update(kwargs)
    return JsonResponse(payload)


# ---------------------------------------------------------------------------
# SCREEN 1 — Prioritization Queue
# ---------------------------------------------------------------------------

@require_http_methods(["GET"])
def api_pendientes(request):
    """
    GET /api/pagos/
    Returns all PENDIENTE invoices sorted by scoring DESC.
    """
    facturas = (
        Factura.objects
        .filter(estado="PENDIENTE")
        .select_related("proveedor")
        .order_by("folio")
    )
    data = sorted(
        [_serialize_factura(f) for f in facturas],
        key=lambda x: x["scoring"],
        reverse=True
    )
    return JsonResponse(data, safe=False)


@require_http_methods(["GET"])
def api_pendientes_por_proveedor(request):
    """
    GET /api/pagos/por-proveedor/
    Retorna las facturas PENDIENTES agrupadas por proveedor, incluyendo la 
    suma total de sus montos, saldos netos y la lista de folios asociados.
    """
    # 1. Traemos todas las facturas pendientes con su proveedor optimizado
    facturas = (
        Factura.objects
        .filter(estado="PENDIENTE")
        .select_related("proveedor")
        .order_by("proveedor__nombre", "folio")
    )

    # 2. Agrupamos en un diccionario usando el ID del proveedor como llave
    proveedores_dict = {}

    for f in facturas:
        prov_id = f.proveedor.id_proveedor

        if prov_id not in proveedores_dict:
            proveedores_dict[prov_id] = {
                "proveedor_id": prov_id,
                "proveedor": f.proveedor.nombre,
                "estrellas": float(f.proveedor.prioridad_estrellas),
                "total_facturas": 0,
                "suma_monto_original": 0.0,
                "suma_total_mxn": 0.0,
                "suma_saldo_neto": 0.0,
                "folios": []
            }

        # 3. Acumulamos los valores de las facturas folidas
        grupo = proveedores_dict[prov_id]
        grupo["total_facturas"] += 1
        grupo["suma_monto_original"] += float(f.monto_original)
        grupo["suma_total_mxn"] += float(f.total)
        grupo["suma_saldo_neto"] += float(f.saldo_neto)
        grupo["folios"].append(f.folio)

    # 4. Convertimos el diccionario a una lista y ordenamos por saldo pendiente de mayor a menor
    data = sorted(
        proveedores_dict.values(),
        key=lambda x: x["suma_saldo_neto"],
        reverse=True
    )

    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def api_autorizar(request, id_factura: int):
    """
    POST /api/facturas/<id>/autorizar/
    Body (JSON): { notas_auditoria: str, abono: float }
    Transitions PENDIENTE → APROBADO.
    """
    try:
        factura = Factura.objects.select_related("proveedor").get(
            pk=id_factura, estado="PENDIENTE"
        )
    except Factura.DoesNotExist:
        return _json_error("Factura no encontrada o no está en estado PENDIENTE.", 404)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return _json_error("JSON inválido en el cuerpo de la solicitud.")

    notas = body.get("notas_auditoria", "")

    # 2. CORRECCIÓN: Convertir el valor del frontend a Decimal de forma segura pasando por string
    try:
        abono = Decimal(str(body.get("abono", 0) or 0))
    except (InvalidOperation, TypeError, ValueError):
        return _json_error("El monto del abono no tiene un formato numérico válido.")

    if abono < 0:
        return _json_error("El abono no puede ser negativo.")

    # 3. CORRECCIÓN: Quitamos float() y comparamos Decimal contra Decimal nativo
    if abono > factura.saldo_neto:
        return _json_error("El abono no puede superar el saldo neto disponible.")

    with transaction.atomic():
        factura.notas_auditoria = notas
        # 4. CORRECCIÓN: Al ser ambos de tipo Decimal, la suma se ejecuta de forma perfecta
        factura.total_abonado = factura.total_abonado + abono
        factura.estado = "APROBADO"
        factura.save()

    return _json_ok(_serialize_factura(factura), message="Factura autorizada correctamente.")


@csrf_exempt
@require_http_methods(["POST"])
def api_revertir_a_pendiente(request, id_factura: int):
    """
    POST /api/facturas/<id>/revertir/
    Transitions APROBADO → PENDIENTE (undo authorization).
    """
    try:
        factura = Factura.objects.select_related("proveedor").get(
            pk=id_factura, estado="APROBADO"
        )
    except Factura.DoesNotExist:
        return _json_error("Factura no encontrada o no está en estado APROBADO.", 404)

    with transaction.atomic():
        factura.estado = "PENDIENTE"
        factura.save()

    return _json_ok(_serialize_factura(factura), message="Factura revertida a PENDIENTE.")


@require_http_methods(["GET"])
def api_facturas_por_proveedor(request, proveedor_id: int):
    """
    GET /api/proveedores/<proveedor_id>/facturas/
    Returns all PENDIENTE invoices for a specific provider, sorted by scoring DESC.
    """
    facturas = (
        Factura.objects
        .filter(proveedor_id=proveedor_id, estado="PENDIENTE")
        .select_related("proveedor")
        .order_by("folio")
    )
    data = sorted(
        [_serialize_factura(f) for f in facturas],
        key=lambda x: x["scoring"],
        reverse=True
    )
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def api_autorizar_lote(request):
    """
    POST /api/facturas/autorizar-lote/
    Body (JSON): {
        "autorizaciones": [
            { "id_factura": int, "abono": float, "notas_auditoria": str },
            ...
        ]
    }
    Transitions PENDIENTE → APROBADO in bulk atomically.
    """
    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return _json_error("JSON inválido en el cuerpo de la solicitud.")

    autorizaciones = body.get("autorizaciones", [])
    if not isinstance(autorizaciones, list):
        return _json_error("El campo 'autorizaciones' debe ser una lista.")

    if not autorizaciones:
        return _json_error("No se enviaron autorizaciones en el lote.")

    factura_ids = [item.get("id_factura") for item in autorizaciones if item.get("id_factura")]
    
    # Prefetch/verify facturas
    facturas_dict = {
        f.id_factura: f for f in 
        Factura.objects.filter(pk__in=factura_ids, estado="PENDIENTE").select_related("proveedor")
    }

    # Validate all items before writing anything
    to_update = []
    errors = []

    for idx, item in enumerate(autorizaciones):
        id_factura = item.get("id_factura")
        if not id_factura:
            errors.append(f"Elemento en índice {idx}: falta 'id_factura'.")
            continue

        factura = facturas_dict.get(id_factura)
        if not factura:
            errors.append(f"Factura ID {id_factura} no encontrada o no está en estado PENDIENTE.")
            continue

        try:
            abono = Decimal(str(item.get("abono", 0) or 0))
        except (InvalidOperation, TypeError, ValueError):
            errors.append(f"Factura {factura.folio}: El monto del abono no es válido.")
            continue

        if abono < 0:
            errors.append(f"Factura {factura.folio}: El abono no puede ser negativo.")
            continue

        if abono > factura.saldo_neto:
            errors.append(f"Factura {factura.folio}: El abono ({abono}) supera el saldo neto disponible ({factura.saldo_neto}).")
            continue

        to_update.append((factura, abono, item.get("notas_auditoria", "")))

    if errors:
        return _json_error("Error de validación en el lote:\n" + "\n".join(errors))

    # Apply changes atomically
    with transaction.atomic():
        serialized_results = []
        for factura, abono, notas in to_update:
            factura.notas_auditoria = notas
            factura.total_abonado = factura.total_abonado + abono
            factura.estado = "APROBADO"
            factura.save()
            serialized_results.append(_serialize_factura(factura))

    return _json_ok(serialized_results, message=f"{len(to_update)} facturas autorizadas correctamente.")


# ---------------------------------------------------------------------------
# SCREEN 2 — Tab A: Confirmation Tray
# ---------------------------------------------------------------------------

@require_http_methods(["GET"])
def api_aprobados(request):
    """
    GET /api/aprobados/
    Returns all APROBADO invoices.
    """
    facturas = (
        Factura.objects
        .filter(estado="APROBADO")
        .select_related("proveedor")
        .order_by("fecha_registro")
    )
    data = [_serialize_factura(f) for f in facturas]
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def api_confirmar(request, id_factura: int):
    """
    POST /api/facturas/<id>/confirmar/
    Transitions APROBADO → CONFIRMADO (locked for export).
    """
    try:
        factura = Factura.objects.select_related("proveedor").get(
            pk=id_factura, estado="APROBADO"
        )
    except Factura.DoesNotExist:
        return _json_error("Factura no encontrada o no está en estado APROBADO.", 404)

    with transaction.atomic():
        factura.estado = "CONFIRMADO"
        factura.save()

    return _json_ok(_serialize_factura(factura), message="Factura confirmada para pago.")


# ---------------------------------------------------------------------------
# SCREEN 2 — Tab B: Export & Reconciliation
# ---------------------------------------------------------------------------

@require_http_methods(["GET"])
def api_confirmados(request):
    """
    GET /api/confirmados/
    Returns all CONFIRMADO invoices.
    """
    facturas = (
        Factura.objects
        .filter(estado="CONFIRMADO")
        .select_related("proveedor")
        .order_by("folio")
    )
    data = [_serialize_factura(f) for f in facturas]
    return JsonResponse(data, safe=False)


@require_http_methods(["GET"])
def api_exportar_remesa(request):
    """
    GET /api/exportar-remesa/
    Generates and returns an Excel (.xlsx) file of all CONFIRMADO invoices.
    """
    facturas = (
        Factura.objects
        .filter(estado="CONFIRMADO")
        .select_related("proveedor")
        .order_by("folio")
    )

    if not facturas.exists():
        return _json_error("No hay facturas confirmadas para exportar.", 404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remesa de Pago"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    center_align = Alignment(horizontal="center", vertical="center")
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    columns = [
        ("Folio",              18),
        ("Proveedor",          30),
        ("Tipo de Pago",       14),
        ("Moneda",             10),
        ("Monto Original",     18),
        ("Tipo de Cambio",     14),
        ("Total MXN",          18),
        ("Total Abonado",      16),
        ("Saldo Neto",         16),
        ("Días Vencidos",      14),
        ("Notas de Auditoría", 40),
        ("Fecha Factura",      16),
    ]

    # Header row
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Alt-row fill
    alt_fill = PatternFill("solid", fgColor="F9FAFB")

    for row_idx, f in enumerate(facturas, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        row_data = [
            f.folio,
            f.proveedor.nombre,
            f.get_tipo_pago_display(),
            f.moneda,
            float(f.monto_original),
            float(f.tipo_cambio),
            float(f.monto_mxn),
            float(f.total_abonado),
            float(f.saldo_neto),
            f.dias_vencidos,
            f.notas_auditoria or "",
            f.fecha_factura.isoformat() if f.fecha_factura else "",
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = fill
            if col_idx in (5, 7, 8, 9):   # money columns
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (6, 10):       # numeric
                cell.alignment = Alignment(horizontal="center")

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"remesa_{TODAY.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
@require_http_methods(["POST"])
def api_conciliar(request):
    """
    POST /api/conciliar/
    Uploads a bank return Excel. Matches folios and transitions CONFIRMADO → PAGADO.
    Expected Excel columns: 'Folio' (required), rest optional.
    Returns a summary: matched, not_found, already_paid.
    """
    if "archivo" not in request.FILES:
        return _json_error("No se encontró el archivo en la solicitud.")

    archivo = request.FILES["archivo"]
    filename = archivo.name.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return _json_error("El archivo debe ser .xlsx o .xls.")

    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    except Exception as e:
        return _json_error(f"No se pudo leer el archivo Excel: {str(e)}")

    ws = wb.active

    # Detect header row and find 'Folio' column
    headers = {}
    folio_col = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for col_idx, cell_val in enumerate(row, 1):
            if cell_val and str(cell_val).strip().lower() == "folio":
                folio_col = col_idx
                # Capture all headers from this row
                for c2_idx, h in enumerate(row, 1):
                    if h:
                        headers[str(h).strip()] = c2_idx
                break
        if folio_col:
            break

    if folio_col is None:
        return _json_error(
            "El archivo no contiene una columna 'Folio'. "
            "Asegúrate de que el encabezado sea exactamente 'Folio'."
        )

    # Find the header row number
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for col_idx, cell_val in enumerate(row, 1):
            if cell_val and str(cell_val).strip().lower() == "folio" and col_idx == folio_col:
                header_row = list(row)
                break
        if header_row:
            break

    # Collect folios from Excel
    folios_in_file = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[folio_col - 1]
        if val:
            folios_in_file.append(str(val).strip())

    if not folios_in_file:
        return _json_error("El archivo no contiene folios en las filas de datos.")

    # Process matches
    matched = []
    not_found = []
    already_paid = []
    wrong_state = []

    with transaction.atomic():
        for folio in folios_in_file:
            try:
                factura = Factura.objects.get(folio=folio)
                if factura.estado == "PAGADO":
                    already_paid.append(folio)
                elif factura.estado == "CONFIRMADO":
                    factura.estado = "PAGADO"
                    factura.save()
                    matched.append(folio)
                else:
                    wrong_state.append(
                        {"folio": folio, "estado": factura.estado})
            except Factura.DoesNotExist:
                not_found.append(folio)

    return _json_ok(
        message=f"{len(matched)} facturas marcadas como PAGADO.",
        matched=matched,
        not_found=not_found,
        already_paid=already_paid,
        wrong_state=wrong_state,
    )


# ---------------------------------------------------------------------------
# Excel Upload (Bulk import of new invoices)
# ---------------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def api_cargar_excel(request):
    """
    POST /api/cargas/excel/
    Bulk-imports invoices from an Excel file.
    Expected columns: Folio, Proveedor, Tipo Pago, Moneda, Monto Original,
                      Tipo Cambio, Total, Fecha Factura, Estrellas
    """
    if "archivo" not in request.FILES:
        return _json_error("No se encontró el archivo.")

    archivo = request.FILES["archivo"]
    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    except Exception as e:
        return _json_error(f"No se pudo leer el archivo: {str(e)}")

    ws = wb.active

    # Map headers
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(v for v in row):
            header_row = [str(v).strip().lower() if v else "" for v in row]
            break

    if not header_row:
        return _json_error("El archivo no contiene encabezados válidos.")

    def col(name):
        try:
            return header_row.index(name)
        except ValueError:
            return None

    FIELD_MAP = {
        "folio":        col("folio"),
        "proveedor":    col("proveedor"),
        "tipo_pago":    col("tipo pago"),
        "moneda":       col("moneda"),
        "monto_orig":   col("monto original"),
        "tipo_cambio":  col("tipo cambio"),
        "total":        col("total"),
        "fecha":        col("fecha factura"),
        "estrellas":    col("estrellas"),
    }

    if FIELD_MAP["folio"] is None or FIELD_MAP["proveedor"] is None:
        return _json_error("El archivo debe contener al menos las columnas 'Folio' y 'Proveedor'.")

    created = []
    errors = []

    with transaction.atomic():
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                folio = str(row[FIELD_MAP["folio"]]).strip(
                ) if row[FIELD_MAP["folio"]] else None
                if not folio:
                    continue  # skip empty rows

                nombre_prov = str(row[FIELD_MAP["proveedor"]]).strip(
                ) if row[FIELD_MAP["proveedor"]] else "Sin nombre"
                estrellas = float(
                    row[FIELD_MAP["estrellas"]]) if FIELD_MAP["estrellas"] is not None and row[FIELD_MAP["estrellas"]] else 3.0
                estrellas = max(1.0, min(5.0, estrellas))

                proveedor, _ = Proveedor.objects.get_or_create(
                    nombre=nombre_prov,
                    defaults={"prioridad_estrellas": estrellas}
                )

                # Determine dates
                fecha_val = row[FIELD_MAP["fecha"]
                                ] if FIELD_MAP["fecha"] is not None else None
                if hasattr(fecha_val, "date"):
                    fecha_val = fecha_val.date()
                elif isinstance(fecha_val, str):
                    from datetime import datetime
                    try:
                        fecha_val = datetime.strptime(
                            fecha_val, "%Y-%m-%d").date()
                    except ValueError:
                        fecha_val = None

                monto_orig = float(
                    row[FIELD_MAP["monto_orig"]]) if FIELD_MAP["monto_orig"] is not None and row[FIELD_MAP["monto_orig"]] else 0.0
                tipo_cambio = float(
                    row[FIELD_MAP["tipo_cambio"]]) if FIELD_MAP["tipo_cambio"] is not None and row[FIELD_MAP["tipo_cambio"]] else 1.0
                total_val = float(
                    row[FIELD_MAP["total"]]) if FIELD_MAP["total"] is not None and row[FIELD_MAP["total"]] else monto_orig * tipo_cambio
                tipo_pago = str(row[FIELD_MAP["tipo_pago"]]).upper(
                ) if FIELD_MAP["tipo_pago"] is not None and row[FIELD_MAP["tipo_pago"]] else "CONTADO"
                if tipo_pago not in ("CONTADO", "CREDITO"):
                    tipo_pago = "CONTADO"
                moneda = str(row[FIELD_MAP["moneda"]]).upper(
                ) if FIELD_MAP["moneda"] is not None and row[FIELD_MAP["moneda"]] else "MXN"

                Factura.objects.update_or_create(
                    folio=folio,
                    defaults={
                        "proveedor":      proveedor,
                        "tipo_pago":      tipo_pago,
                        "moneda":         moneda,
                        "monto_original": monto_orig,
                        "monto_mxn":      total_val,
                        "tipo_cambio":    tipo_cambio,
                        "total":          total_val,
                        "total_abonado":  0,
                        "fecha_factura":  fecha_val,
                        "estado":         "PENDIENTE",
                    }
                )
                created.append(folio)

            except Exception as e:
                errors.append({"fila": row_num, "error": str(e)})

    return _json_ok(
        message=f"{len(created)} facturas importadas correctamente.",
        created=created,
        errors=errors,
    )
